#!/usr/bin/python3
#RenameColumns.py
#created by Jeff Weltman
#https://github.com/jeffweltman/python_scripts
#v1.0 8/29/2017

def main():
    import os
    import openpyxl
    target_dir = ('Documents') # set target directory
    for file in os.listdir(target_dir):
        filename = os.fsdecode(file)
        if filename.endswith('.xlsx'): #looking only at Excel docs
            book = filename.rstrip('.xlsx') # this strips the extension
            print(book)
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, max_col=75, max_row=1): #setting only the first row to get the headers
                for cellObj in row:
                    print(cellObj.value)
                    if cellObj.value == 'sem_exam': #search term for inaccurate column name
                        cellObj.value = 'n_sem_exam' #replace search term with this target column name
                    elif cellObj.value == 'sem_final':
                        cellObj.value = 'n_sem_final'
                    elif not cellObj.value: break #break after the first null column name if max_col number above exceeds existing column number
                    else: continue
                wb.save(filename)
        
if __name__ == "__main__": main()
