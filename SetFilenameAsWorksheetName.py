#!/usr/bin/python3
#SetFilenameAsWorksheetName.py
#created by Jeff Weltman
#https://github.com/jeffweltman/python_scripts
#v3.0 10/30/2017
#Requires openpyxl package

def main():
    import os
    import openpyxl
    target_dir = ('Documents') # set target directory
    for file in os.listdir(target_dir):
        filename = os.fsdecode(file)
        if filename.endswith('.xlsx'): #looking only at Excel docs
            book = filename.rstrip('.xlsx') # this strips the extension
            fullpath = target_dir + "/" + filename
            #print(book) 
            badname = 'Worksheet for {} was not correct.'.format(book)
            goodname = 'Worksheet for {} is correct.'.format(book)
            from openpyxl import load_workbook
            wb = load_workbook(fullpath)
            ws = wb.active
            #print(ws)
            #print(wb.sheetnames)
            for sheet in wb:
                if ws.title != book:
                    ws.title = book
                    print(badname)
                else:
                  print(goodname)
                  continue
                wb.save(fullpath)

if __name__ == "__main__": main()
